# SPDX-License-Identifier: Apache-2.0
"""Parseurs GeoJSON, CSV+lat/lon (SP-6a) et GeoPackage/Shapefile zippé
(SP-6b, via pyogrio — wheels manylinux, GDAL/GEOS/PROJ embarqués, aucun
paquet système requis). Chaque parseur produit un flux (géométrie shapely,
propriétés) ; toute ligne/feature/entité invalide lève IngestionParseError
immédiatement (fail-fast) — pas d'import partiel silencieux."""

import csv
import datetime
import io
import json
import math
import tempfile
import warnings
from collections.abc import Iterator
from contextlib import contextmanager
from dataclasses import dataclass

import geopandas as gpd
import numpy as np
import pyogrio
import pyproj
import shapely
from openpyxl import load_workbook
from pyogrio.errors import DataLayerError, DataSourceError
from pyproj.exceptions import ProjError
from shapely.errors import ShapelyError
from shapely.geometry import Point, shape
from shapely.geometry.base import BaseGeometry
from shapely.ops import transform as shapely_transform


class IngestionParseError(Exception):
    """Message affiché tel quel comme ingestion_jobs.error_message."""


_LAT_NAMES = {"lat", "latitude", "y"}
_LON_NAMES = {"lon", "lng", "longitude", "x"}
_WGS84 = pyproj.CRS.from_epsg(4326)
_OGR_ERRORS = (DataSourceError, DataLayerError)


def detect_lat_lon_fields(fieldnames: list[str]) -> tuple[str, str] | None:
    by_lower = {name.lower(): name for name in fieldnames}
    lat = next((by_lower[n] for n in _LAT_NAMES if n in by_lower), None)
    lon = next((by_lower[n] for n in _LON_NAMES if n in by_lower), None)
    if lat is None or lon is None:
        return None
    return lat, lon


def parse_geojson(content: bytes) -> Iterator[tuple[BaseGeometry, dict]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise IngestionParseError("encodage invalide, attendu UTF-8") from exc
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        raise IngestionParseError(f"JSON invalide : {exc}") from exc
    if not isinstance(data, dict) or data.get("type") != "FeatureCollection":
        raise IngestionParseError("le GeoJSON doit être une FeatureCollection")
    features = data.get("features", [])
    if not isinstance(features, list):
        raise IngestionParseError("le GeoJSON doit être une FeatureCollection")
    for i, feature in enumerate(features):
        if not isinstance(feature, dict):
            raise IngestionParseError(f"feature {i} : entrée invalide")
        geometry = feature.get("geometry")
        if geometry is None:
            raise IngestionParseError(f"feature {i} : géométrie manquante")
        try:
            geom = shape(geometry)
        except (ValueError, AttributeError, KeyError, TypeError, ShapelyError) as exc:
            raise IngestionParseError(f"feature {i} : géométrie invalide ({exc})") from exc
        if not geom.is_valid:
            raise IngestionParseError(f"feature {i} : géométrie invalide")
        properties = feature.get("properties")
        if properties is not None and not isinstance(properties, dict):
            raise IngestionParseError(f"feature {i} : properties invalide")
        yield geom, dict(properties or {})


def parse_csv_latlon(
    content: bytes,
    lat_field: str | None,
    lon_field: str | None,
) -> Iterator[tuple[BaseGeometry, dict]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise IngestionParseError("encodage invalide, attendu UTF-8") from exc
    reader = csv.DictReader(io.StringIO(text))
    try:
        fieldnames = reader.fieldnames or []
    except csv.Error as exc:
        raise IngestionParseError("en-tête CSV invalide ou mal formé") from exc
    if lat_field is None or lon_field is None:
        detected = detect_lat_lon_fields(fieldnames)
        if detected is None:
            raise IngestionParseError(
                "colonnes lat/lon introuvables automatiquement — précisez-les"
            )
        lat_field, lon_field = detected
    if lat_field not in fieldnames or lon_field not in fieldnames:
        raise IngestionParseError(f"colonnes '{lat_field}'/'{lon_field}' absentes du CSV")
    i = 0
    row_iter = iter(reader)
    while True:
        try:
            row = next(row_iter)
        except StopIteration:
            break
        except csv.Error as exc:
            raise IngestionParseError(
                f"ligne {i + 1} : champ CSV trop volumineux ou mal formé"
            ) from exc
        i += 1
        try:
            lat = float(row[lat_field])
            lon = float(row[lon_field])
        except (TypeError, ValueError):
            raise IngestionParseError(
                f"ligne {i} : lat/lon invalide ('{row.get(lat_field)}', '{row.get(lon_field)}')"
            ) from None
        properties = {k: v for k, v in row.items() if k not in (lat_field, lon_field)}
        yield Point(lon, lat), properties


def _xlsx_cell_value(value):
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    return value


def parse_xlsx_latlon(
    content: bytes,
    lat_field: str | None,
    lon_field: str | None,
) -> Iterator[tuple[BaseGeometry, dict]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise IngestionParseError("classeur XLSX vide") from None
    fieldnames = [str(name) if name is not None else "" for name in header_row]
    if lat_field is None or lon_field is None:
        detected = detect_lat_lon_fields(fieldnames)
        if detected is None:
            raise IngestionParseError(
                "colonnes lat/lon introuvables automatiquement — précisez-les"
            )
        lat_field, lon_field = detected
    if lat_field not in fieldnames or lon_field not in fieldnames:
        raise IngestionParseError(f"colonnes '{lat_field}'/'{lon_field}' absentes du XLSX")
    lat_idx = fieldnames.index(lat_field)
    lon_idx = fieldnames.index(lon_field)
    for i, row in enumerate(rows_iter, start=1):
        raw_lat = row[lat_idx] if lat_idx < len(row) else None
        raw_lon = row[lon_idx] if lon_idx < len(row) else None
        try:
            lat = float(raw_lat)
            lon = float(raw_lon)
        except (TypeError, ValueError):
            raise IngestionParseError(
                f"ligne {i} : lat/lon invalide ('{raw_lat}', '{raw_lon}')"
            ) from None
        properties = {
            name: _xlsx_cell_value(row[j] if j < len(row) else None)
            for j, name in enumerate(fieldnames)
            if j not in (lat_idx, lon_idx)
        }
        yield Point(lon, lat), properties


@dataclass
class LayerInfo:
    name: str
    feature_count: int
    geometry_type: str


@contextmanager
def _temp_file(content: bytes, suffix: str) -> Iterator[str]:
    with tempfile.NamedTemporaryFile(suffix=suffix) as tmp:
        tmp.write(content)
        tmp.flush()
        yield tmp.name


def _crs_transform(crs: str | None):
    try:
        src = pyproj.CRS.from_user_input(crs)
    except ProjError as exc:
        raise IngestionParseError(f"CRS manquant ou non reconnu : {crs!r}") from exc
    if src == _WGS84:
        return None
    try:
        transformer = pyproj.Transformer.from_crs(src, _WGS84, always_xy=True)
    except ProjError as exc:
        raise IngestionParseError(f"CRS non transformable vers WGS84 : {crs!r}") from exc
    return transformer.transform


def _native_value(value):
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def _read_features(path: str, layer_name: str | None) -> Iterator[tuple[BaseGeometry, dict]]:
    try:
        raw_layers = pyogrio.list_layers(path)
    except _OGR_ERRORS as exc:
        raise IngestionParseError(f"fichier illisible : {exc}") from exc
    available = [str(name) for name, _geom_type in raw_layers]
    if layer_name is None:
        if len(available) != 1:
            raise IngestionParseError(
                f"plusieurs couches disponibles ({', '.join(available)}) — précisez layerName"
            )
        layer_name = available[0]
    elif layer_name not in available:
        raise IngestionParseError(
            f"couche '{layer_name}' introuvable — couches disponibles : {', '.join(available)}"
        )
    try:
        # pyogrio/numpy émet un DeprecationWarning interne ("generic unit for
        # NumPy timedelta") sur des champs datetime NaT (rencontré en
        # pratique sur les champs begin/end du schéma KML par défaut) —
        # vérifié par exécution réelle (SP-56), quirk de la bibliothèque
        # tierce, sans rapport avec la validité des données lues.
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message=".*generic.*unit for NumPy timedelta.*",
                category=DeprecationWarning,
            )
            meta, _index, geometry, field_data = pyogrio.raw.read(
                path, layer=layer_name, force_2d=True
            )
    except _OGR_ERRORS as exc:
        raise IngestionParseError(f"couche '{layer_name}' illisible : {exc}") from exc

    transform = _crs_transform(meta["crs"])
    fields = list(meta["fields"])

    for i, wkb in enumerate(geometry):
        if wkb is None:
            raise IngestionParseError(f"entité {i} : géométrie manquante")
        try:
            geom = shapely.from_wkb(wkb)
        except ShapelyError as exc:
            raise IngestionParseError(f"entité {i} : géométrie invalide ({exc})") from exc
        if transform is not None:
            geom = shapely_transform(transform, geom)
        if not geom.is_valid:
            raise IngestionParseError(f"entité {i} : géométrie invalide")
        properties = {field: _native_value(field_data[j][i]) for j, field in enumerate(fields)}
        yield geom, properties


def parse_gpkg(
    content: bytes,
    layer_name: str | None = None,
) -> Iterator[tuple[BaseGeometry, dict]]:
    with _temp_file(content, ".gpkg") as path:
        yield from _read_features(path, layer_name)


def parse_shapefile_zip(
    content: bytes,
    layer_name: str | None = None,
) -> Iterator[tuple[BaseGeometry, dict]]:
    with _temp_file(content, ".zip") as path:
        yield from _read_features(f"/vsizip/{path}", layer_name)


# Le driver KML de GDAL impose un schéma de champs fixe sur TOUT Placemark,
# y compris un champ nommé "id" (l'attribut XML id="..." du Placemark, vide
# sinon) — vérifié par exécution réelle : même un KML minimal à un seul
# Placemark sans schéma personnalisé le produit. Ce nom entre en collision
# avec la colonne "id" (clé primaire serial) que run_import pose sur toute
# table importée ; sans renommage, TOUT import KML échouerait à la création
# de table ("column id specified more than once"), pas seulement un cas
# limite de nommage utilisateur. Renommé plutôt que supprimé pour ne pas
# perdre l'attribut id du Placemark quand il est renseigné.
_KML_RESERVED_PROPERTY_NAMES = {"id", "tenant_id", "geom"}


def _rename_kml_reserved_properties(props: dict) -> dict:
    return {
        (f"kml_{key}" if key in _KML_RESERVED_PROPERTY_NAMES else key): value
        for key, value in props.items()
    }


def parse_kml(
    content: bytes,
    layer_name: str | None = None,
) -> Iterator[tuple[BaseGeometry, dict]]:
    # Un .kmz est un zip contenant un doc.kml, mais GDAL/pyogrio le détecte
    # et le lit DIRECTEMENT sur l'extension .kmz elle-même — PAS de préfixe
    # /vsizip/ ici, contrairement à parse_shapefile_zip (.zip) : ce préfixe
    # casserait la lecture ("n'est pas un fichier kmz valide"), vérifié par
    # exécution réelle avant d'écrire ce code (spec SP-56 §Contexte). Le
    # suffixe temporaire distingue .kml/.kmz uniquement pour que GDAL
    # sélectionne le bon driver depuis l'extension.
    suffix = ".kmz" if _looks_like_zip(content) else ".kml"
    with _temp_file(content, suffix) as path:
        for geom, props in _read_features(path, layer_name):
            yield geom, _rename_kml_reserved_properties(props)


def _looks_like_zip(content: bytes) -> bool:
    return content[:2] == b"PK"


def parse_geoparquet(content: bytes) -> Iterator[tuple[BaseGeometry, dict]]:
    # PAS pyogrio : pyogrio.list_drivers()["Parquet"] vaut None dans ce build
    # (aucun driver OGR Parquet) — vérifié par exécution réelle (spec SP-56
    # §3). geopandas.read_parquet() lit correctement un GeoParquet 1.0,
    # y compris celui produit par app.cdc.parquet_writer.write_geoparquet
    # (SP-11). Un seul fichier, pas de concept de couches multiples ici :
    # ce format ne passe jamais par list_layers()/l'étape d'inspection.
    with _temp_file(content, ".parquet") as path:
        gdf = gpd.read_parquet(path)
        if gdf.crs is not None and gdf.crs.to_epsg() != 4326:
            gdf = gdf.to_crs(epsg=4326)
        geom_col = gdf.geometry.name
        for i, row in gdf.iterrows():
            geom = row[geom_col]
            # Une géométrie manquante revient de geopandas.read_parquet en
            # NaN (float), pas None — vérifié par exécution réelle (écart
            # avec le pseudo-code de la spec SP-56 §3.1, corrigé ici).
            if geom is None or (isinstance(geom, float) and math.isnan(geom)):
                raise IngestionParseError(f"entité {i} : géométrie manquante")
            props = {k: _native_value(v) for k, v in row.items() if k != geom_col}
            yield geom, props


def list_layers(content: bytes, filename: str) -> list[LayerInfo]:
    lower = filename.lower()
    if lower.endswith(".gpkg"):
        suffix, wrap = ".gpkg", (lambda p: p)
    elif lower.endswith(".zip"):
        suffix, wrap = ".zip", (lambda p: f"/vsizip/{p}")
    elif lower.endswith((".kml", ".kmz")):
        # Identité : PAS le wrap /vsizip/ de la branche .zip ci-dessus, cf.
        # parse_kml — un .kmz se lit tel quel.
        suffix, wrap = (lower[lower.rfind(".") :], lambda p: p)
    else:
        raise ValueError(f"format non concerné par l'inspection : {filename}")
    with _temp_file(content, suffix) as tmp_path:
        path = wrap(tmp_path)
        try:
            raw_layers = pyogrio.list_layers(path)
        except _OGR_ERRORS as exc:
            raise IngestionParseError(f"fichier illisible : {exc}") from exc
        layers = []
        for name, _geom_type in raw_layers:
            try:
                info = pyogrio.read_info(path, layer=name)
            except _OGR_ERRORS as exc:
                raise IngestionParseError(f"couche '{name}' illisible : {exc}") from exc
            layers.append(
                LayerInfo(
                    name=str(name),
                    feature_count=int(info["features"]),
                    geometry_type=str(info["geometry_type"] or "Unknown"),
                )
            )
        if not layers:
            raise IngestionParseError("aucune couche trouvée dans le fichier")
        return layers
